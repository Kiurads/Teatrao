"""
GUI version for extracting data from multiple Excel files into a Bordereau report.
"""
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from datetime import datetime
from pathlib import Path
from typing import Union
import threading

import openpyxl
from openpyxl.utils import column_index_from_string


# Constants
SOURCE_SHEET_NAME = "Folha1"
OUTPUT_SHEET_NAME = "Bordereaux_Geral"

ITEMS_TO_EXTRACT_STATIC = [
    "F1", "F5", None, None, "F7", "F3", "F15", "F13", "F9", "F11", "F17"
]

SUMMARY_COLUMNS = [
    (6, "F"),  # F column (column 6)
    (8, "H"),  # H column (column 8)
    (10, "J"), # J column (column 10)
    (4, "D"),  # D column (column 4)
    (5, "E"),  # E column (column 5)
    (2, "B", 5),  # B column (column 2) - for observations (5 rows below Convites)
]

MONTHS = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
]

WEEKDAYS = [
    "Segunda-feira", "Terça-feira", "Quarta-feira",
    "Quinta-feira", "Sexta-feira", "Sábado", "Domingo"
]

STATIC_HEADER_COLUMNS = [
    "Nº Registo", "Data", "Mês", "Dia da Semana", "Hora", "Nome do Espetáculo",
    "Local", "Atividade", "Classificação Etária", "Evento", "Lotação Máxima"
]

SUMMARY_HEADER_COLUMNS = [
    "Total Bilheteira", "Total Postos TL", "Total Internet", "Total de Bilhetes", "Valor Total", "Observação"
]


def column_to_number(column_letter: str) -> int:
    """Convert Excel column letter to column number."""
    return column_index_from_string(column_letter)


def extract_cell_value(sheet, cell_ref: str) -> Union[str, int, float, datetime, None]:
    """Extract value from a cell reference."""
    column_letter = ''.join(filter(str.isalpha, cell_ref))
    row_number = int(''.join(filter(str.isdigit, cell_ref)))
    return sheet.cell(row_number, column_to_number(column_letter)).internal_value


def format_date_value(value: datetime) -> str:
    """Format datetime value to string."""
    return value.strftime("%d-%m-%Y")


class BordereauGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Gerador de Proposta Bordereau")
        self.root.geometry("800x500")
        self.root.resizable(True, True)
        
        self.input_folder = tk.StringVar(value=os.getcwd())
        self.output_file = tk.StringVar(value="Proposta_Bordereau.xlsx")
        self.processing = False
        
        self.create_widgets()
        
    def create_widgets(self):
        """Create all GUI widgets."""
        # Title
        title_frame = ttk.Frame(self.root, padding="5")
        title_frame.pack(fill=tk.X)
        
        title_label = ttk.Label(
            title_frame, 
            text="Gerador de Proposta Bordereau", 
            font=("Segoe UI", 14, "bold")
        )
        title_label.pack()
        
        subtitle = ttk.Label(
            title_frame,
            text="Extrair dados de múltiplos ficheiros Excel para um relatório consolidado",
            font=("Segoe UI", 8)
        )
        subtitle.pack()
        
        # Separator
        ttk.Separator(self.root, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=3)
        
        # Input folder selection
        folder_frame = ttk.LabelFrame(self.root, text="Pasta de Origem", padding="5")
        folder_frame.pack(fill=tk.X, padx=10, pady=2)
        
        folder_entry_frame = ttk.Frame(folder_frame)
        folder_entry_frame.pack(fill=tk.X)
        
        ttk.Entry(
            folder_entry_frame, 
            textvariable=self.input_folder, 
            state="readonly"
        ).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        ttk.Button(
            folder_entry_frame, 
            text="Procurar...", 
            command=self.browse_folder
        ).pack(side=tk.RIGHT)
        
        # Output file name
        output_frame = ttk.LabelFrame(self.root, text="Ficheiro de Saída", padding="5")
        output_frame.pack(fill=tk.X, padx=10, pady=2)
        
        ttk.Entry(
            output_frame, 
            textvariable=self.output_file
        ).pack(fill=tk.X)
        
        # Progress bar
        progress_frame = ttk.Frame(self.root, padding="5")
        progress_frame.pack(fill=tk.X, padx=10, pady=2)
        
        self.progress = ttk.Progressbar(
            progress_frame, 
            mode='indeterminate'
        )
        self.progress.pack(fill=tk.X)
        
        # Buttons
        button_frame = ttk.Frame(self.root, padding="5")
        button_frame.pack(fill=tk.X, padx=10, pady=2)
        
        self.process_btn = ttk.Button(
            button_frame, 
            text="Processar Ficheiros", 
            command=self.start_processing,
            style="Accent.TButton"
        )
        self.process_btn.pack(side=tk.LEFT, padx=3)
        
        ttk.Button(
            button_frame, 
            text="Limpar Log", 
            command=self.clear_log
        ).pack(side=tk.LEFT, padx=3)
        
        # Log output
        log_frame = ttk.LabelFrame(self.root, text="Log de Processamento", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=3)
        
        self.log_text = scrolledtext.ScrolledText(
            log_frame, 
            wrap=tk.WORD,
            font=("Consolas", 8)
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # Status bar
        self.status_var = tk.StringVar(value="Pronto")
        status_bar = ttk.Label(
            self.root, 
            textvariable=self.status_var, 
            relief=tk.SUNKEN, 
            anchor=tk.W
        )
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
    def browse_folder(self):
        """Open folder browser dialog."""
        folder = filedialog.askdirectory(
            initialdir=self.input_folder.get(),
            title="Selecione a pasta com os ficheiros Excel"
        )
        if folder:
            self.input_folder.set(folder)
            self.log(f"Pasta selecionada: {folder}")
    
    def log(self, message: str):
        """Add message to log text area."""
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def clear_log(self):
        """Clear the log text area."""
        self.log_text.delete(1.0, tk.END)
    
    def update_status(self, message: str):
        """Update status bar."""
        self.status_var.set(message)
        self.root.update_idletasks()
    
    def start_processing(self):
        """Start processing in a separate thread."""
        if self.processing:
            messagebox.showwarning("Aviso", "Processamento já em curso!")
            return
        
        # Validate inputs
        if not Path(self.input_folder.get()).exists():
            messagebox.showerror("Erro", "A pasta selecionada não existe!")
            return
        
        if not self.output_file.get().strip():
            messagebox.showerror("Erro", "Especifique um nome para o ficheiro de saída!")
            return
        
        # Start processing in background thread
        self.processing = True
        self.process_btn.config(state="disabled")
        self.progress.start(10)
        
        thread = threading.Thread(target=self.process_files, daemon=True)
        thread.start()
    
    def process_files(self):
        """Process all Excel files."""
        try:
            self.log("=" * 60)
            self.log("Iniciando processamento...")
            self.log("=" * 60)
            self.update_status("A processar ficheiros...")
            
            input_path = Path(self.input_folder.get())
            output_path = input_path / self.output_file.get()
            
            # Remove existing output file
            try:
                output_path.unlink()
                self.log(f"Ficheiro {output_path.name} removido (vai ser reposto).")
            except FileNotFoundError:
                self.log(f"Ficheiro {output_path.name} vai ser criado.")
            except PermissionError:
                self.log(f"ERRO: Não tem permissões para repor {output_path.name}")
                self.log("Certifique-se de que o ficheiro não está aberto.")
                messagebox.showerror(
                    "Erro de Permissão",
                    f"Não foi possível sobrescrever {output_path.name}.\n"
                    "Certifique-se de que o ficheiro não está aberto."
                )
                return
            
            # Create output workbook
            output_wb = openpyxl.Workbook()
            output_ws = output_wb.active
            output_ws.title = OUTPUT_SHEET_NAME
            
            # Find Excel files
            excel_files = [
                f for f in input_path.iterdir()
                if f.suffix == ".xlsx" and not f.name.startswith("~$") and f.name != output_path.name
            ]
            
            if not excel_files:
                self.log("\nNenhum ficheiro Excel encontrado para processar.")
                messagebox.showinfo("Informação", "Nenhum ficheiro Excel encontrado na pasta selecionada.")
                return
            
            # Generate headers from the first file
            try:
                header_columns = self.generate_headers(excel_files[0])
                output_ws.append(header_columns)
            except Exception as e:
                self.log(f"ERRO ao gerar headers: {e}")
                messagebox.showerror("Erro", f"Não foi possível gerar os headers:\n\n{e}")
                return
            
            self.log(f"\nEncontrados {len(excel_files)} ficheiro(s) para processar.\n")
            
            # Process each file
            output_row = 2
            processed_count = 0
            error_count = 0
            
            for excel_file in sorted(excel_files):
                self.log(f"A processar: {excel_file.name}")
                self.update_status(f"A processar: {excel_file.name}")
                
                try:
                    self.process_excel_file(excel_file, output_ws, output_row)
                    output_row += 1
                    processed_count += 1
                except Exception as e:
                    self.log(f"  ERRO: {e}")
                    error_count += 1
            
            # Save output file
            output_wb.save(output_path)
            
            self.log("\n" + "=" * 60)
            self.log(f"Processamento concluído!")
            self.log(f"Ficheiros processados: {processed_count}")
            if error_count > 0:
                self.log(f"Ficheiros com erros: {error_count}")
            self.log(f"Ficheiro guardado: {output_path.name}")
            self.log("=" * 60)
            
            self.update_status("Processamento concluído com sucesso!")
            
            messagebox.showinfo(
                "Sucesso",
                f"Processamento concluído!\n\n"
                f"Ficheiros processados: {processed_count}\n"
                f"Ficheiros com erros: {error_count}\n\n"
                f"Ficheiro guardado: {output_path.name}"
            )
            
        except Exception as e:
            self.log(f"\nERRO FATAL: {e}")
            self.update_status("Erro durante o processamento")
            messagebox.showerror("Erro Fatal", f"Ocorreu um erro durante o processamento:\n\n{e}")
        
        finally:
            self.processing = False
            self.process_btn.config(state="normal")
            self.progress.stop()
    
    def generate_headers(self, file_path: Path) -> list:
        """Generate header columns dynamically from the first Excel file."""
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        if SOURCE_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"Folha '{SOURCE_SHEET_NAME}' não encontrada")
        
        sheet = workbook[SOURCE_SHEET_NAME]
        
        # Start with static headers
        headers = STATIC_HEADER_COLUMNS.copy()
        
        # Find the row where column B contains "Convites"
        convites_row = None
        for row_num in range(24, 100):
            cell_value = sheet.cell(row_num, 2).value  # Column B is column 2
            if cell_value == "Convites":
                convites_row = row_num
                break
        
        if convites_row is None:
            raise ValueError("Linha 'Convites' não encontrada")
        
        # Extract category names and prices from rows 24 to convites_row (inclusive)
        for row_num in range(24, convites_row + 1):
            category_name = sheet.cell(row_num, 2).value  # Column B
            price_value = sheet.cell(row_num, 3).value    # Column C
            
            if category_name:
                headers.append(category_name)
                headers.append("Valor " + str(category_name))
        
        # Add summary headers
        headers.extend(SUMMARY_HEADER_COLUMNS)
        
        workbook.close()
        return headers
    
    def process_excel_file(self, file_path: Path, output_ws, output_row: int):
        """Process a single Excel file and extract data."""
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        if SOURCE_SHEET_NAME not in workbook.sheetnames:
            self.log(f"  Aviso: Folha '{SOURCE_SHEET_NAME}' não encontrada")
            workbook.close()
            raise ValueError(f"Folha '{SOURCE_SHEET_NAME}' não encontrada")
        
        sheet = workbook[SOURCE_SHEET_NAME]
        
        # Extract static items (header info)
        col_index = 1
        for item in ITEMS_TO_EXTRACT_STATIC:
            if item is None:
                col_index += 1
                continue
            
            value = extract_cell_value(sheet, item)
            
            if isinstance(value, datetime):
                output_ws.cell(output_row, col_index, format_date_value(value))
                output_ws.cell(output_row, 3, MONTHS[value.month - 1])
                output_ws.cell(output_row, 4, WEEKDAYS[value.weekday()])
            else:
                output_ws.cell(output_row, col_index, value)
            col_index += 1
        
        # Find the row where column B contains "Convites"
        convites_row = None
        for row_num in range(24, 100):
            cell_value = sheet.cell(row_num, 2).value  # Column B is column 2
            if cell_value == "Convites":
                convites_row = row_num
                break
        
        if convites_row is None:
            self.log(f"  Aviso: Não foi encontrada a linha 'Convites'")
            workbook.close()
            raise ValueError("Linha 'Convites' não encontrada")
        
        # Extract D and E columns from row 24 to the "Convites" row (inclusive)
        for row_num in range(24, convites_row + 1):
            d_value = sheet.cell(row_num, 4).internal_value  # Column D
            e_value = sheet.cell(row_num, 5).internal_value  # Column E
            
            output_ws.cell(output_row, col_index, d_value)
            col_index += 1
            output_ws.cell(output_row, col_index, e_value)
            col_index += 1
        
        # Calculate summary row (2 rows below Convites)
        summary_row = convites_row + 2
        
        # Extract summary items from columns: F, H, J, D, E, and B (at appropriate rows)
        for summary_col_info in SUMMARY_COLUMNS:
            col_num = summary_col_info[0]
            col_letter = summary_col_info[1]
            
            if len(summary_col_info) > 2:
                # Custom offset for this column (e.g., observations)
                row_offset = summary_col_info[2]
                value = sheet.cell(convites_row + row_offset, col_num).internal_value
            else:
                # Default: use summary_row (2 rows below Convites)
                value = sheet.cell(summary_row, col_num).internal_value
            
            output_ws.cell(output_row, col_index, value)
            col_index += 1
        
        workbook.close()
        self.log(f"  ✓ Processado com sucesso")


def main():
    """Main application entry point."""
    root = tk.Tk()
    app = BordereauGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
