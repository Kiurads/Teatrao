"""
Script to extract data from multiple Excel files and consolidate into a single Bordereau report.
"""
import os
from datetime import datetime
from pathlib import Path
from typing import Union

import openpyxl
from openpyxl.utils import column_index_from_string


def column_to_number(column_letter: str) -> int:
    """
    Convert Excel column letter to column number.
    
    Args:
        column_letter: Column letter (e.g., 'A', 'B', 'AA')
    
    Returns:
        Column number (1-indexed)
    """
    return column_index_from_string(column_letter)



# Constants
OUTPUT_FILE = Path("./Proposta_Bordereau.xlsx")
SOURCE_SHEET_NAME = "Folha1"
OUTPUT_SHEET_NAME = "Bordereaux_Geral"

# Items to look for (cell references)
ITEMS_TO_EXTRACT = [
    "F1", "F5", None, None, "F7", "F3", "F15", "F13", "F9", "F11", "F17", "D24", "E24",
    "D25", "E25", "D26", "E26", "D27", "E27", "D28", "E28", "D29", "E29",
    "D30", "E30", "D31", "E31", "D32", "E32", "D33", "E33", "D34", "E34",
    "D35", "E35", "D36", "E36", "D37", "E37", "D38", "E38", "D39", "E39",
    "D40", "E40", "D41", "E41", "D42", "E42", "D43", "E43", "D44", "E44",
    "D45", "E45", "D46", "E46", "D47", "E47", "D48", "E48", "D49", "F51",
    "H51", "J51", "D51", "E51", "B54"
]

# List of months in Portuguese
MONTHS = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
]

# List of weekdays in Portuguese, starting with Monday
WEEKDAYS = [
    "Segunda-feira", "Terça-feira", "Quarta-feira",
    "Quinta-feira", "Sexta-feira", "Sábado", "Domingo"
]

# Header for output
HEADER_COLUMNS = [
    "Nº Registo", "Data", "Mês", "Dia da Semana", "Hora", "Nome do Espetáculo",
    "Local", "Atividade", "Classificação Etária", "Evento", "Lotação Máxima",
    "Normal", "Valor", "Entidades Protocoladas", "Valor2", "Estudantes",
    "Valor3", "Seniores > 65", "Valor4", "Profissionais das Artes Espetáculo",
    "Valor5", "Pais Classes do Teatrão", "Valor6", "Desempregados", "Valor7",
    "Alunos São Teotónio", "Valor8", "Alunos Classe de Teatro", "Valor9",
    "Colaboradores IPC", "Valor10", "Comunidades Vale das Flores", "Valor11",
    "Grupos a partir 10 Pessoas", "Valor12", "Pessoas Portadoras de Deficiência e/ou S/surdas",
    "Valor13", "Colaboradores ISEC e ESEC; Alunos IPC", "Valor14", "Projeto Pedagógico",
    "Valor15", "MÚSICA", "Valor16", "MÚSICA_Conservatória", "Valor17",
    "Res. Artística 9 Normal", "Valor18", "Res. Artística 9 Estudantes",
    "Valor19", "Res. Artística 9 Classes de Teatro", "Valor20",
    "Outras Situações Pré-Venda", "Valor24", "Outras Situações Venda",
    "Valor25", "Escolas", "Valor21", "Escolas com transporte", "Valor22",
    "Escolas com Protocolo", "Valor23", "Convites", "Total Bilheteira",
    "Total Postos TL", "Total Internet", "Total de Bilhetes", "Valor Total",
    "Observação"
]


def prepare_output_file() -> None:
    """Remove existing output file if it exists."""
    try:
        OUTPUT_FILE.unlink()
        print(f"O ficheiro {OUTPUT_FILE} vai ser reposto.")
    except FileNotFoundError:
        print(f"O ficheiro {OUTPUT_FILE} vai ser criado.")
    except PermissionError:
        print(f"Erro: Não tem permissões para repor {OUTPUT_FILE}.")
        print("Certifique-se de que o ficheiro não está aberto.")
        raise


def extract_cell_value(sheet, cell_ref: str) -> Union[str, int, float, datetime, None]:
    """
    Extract value from a cell reference.
    
    Args:
        sheet: Excel worksheet object
        cell_ref: Cell reference (e.g., 'A1', 'B5')
    
    Returns:
        Cell value
    """
    column_letter = ''.join(filter(str.isalpha, cell_ref))
    row_number = int(''.join(filter(str.isdigit, cell_ref)))
    
    return sheet.cell(row_number, column_to_number(column_letter)).internal_value


def format_date_value(value: datetime) -> str:
    """Format datetime value to string."""
    return value.strftime("%d-%m-%Y")


def process_excel_file(file_path: Path, output_ws, output_row: int) -> None:
    """
    Process a single Excel file and extract data.
    
    Args:
        file_path: Path to the Excel file
        output_ws: Output worksheet object
        output_row: Current row in output worksheet
    """
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        if SOURCE_SHEET_NAME not in workbook.sheetnames:
            print(f"  Aviso: Folha '{SOURCE_SHEET_NAME}' não encontrada em {file_path.name}")
            return
        
        sheet = workbook[SOURCE_SHEET_NAME]
        
        for col_index, item in enumerate(ITEMS_TO_EXTRACT, start=1):
            if item is None:
                continue
            
            value = extract_cell_value(sheet, item)
            
            if isinstance(value, datetime):
                output_ws.cell(output_row, col_index, format_date_value(value))
                # Set month and weekday in columns 3 and 4
                output_ws.cell(output_row, 3, MONTHS[value.month - 1])
                output_ws.cell(output_row, 4, WEEKDAYS[value.weekday()])
            else:
                output_ws.cell(output_row, col_index, value)
        
        workbook.close()
        
    except Exception as e:
        print(f"  Erro ao processar {file_path.name}: {e}")


def main() -> None:
    """Main execution function."""
    print("=" * 60)
    print("Extração de dados para Bordereau")
    print("=" * 60)
    
    # Prepare output file
    prepare_output_file()
    
    # Create output workbook
    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active
    output_ws.title = OUTPUT_SHEET_NAME
    
    # Add header
    output_ws.append(HEADER_COLUMNS)
    
    # Process files
    current_dir = Path("./")
    excel_files = [
        f for f in current_dir.iterdir() 
        if f.suffix == ".xlsx" and not f.name.startswith("~$") and f.name != OUTPUT_FILE.name
    ]
    
    if not excel_files:
        print("\nNenhum ficheiro Excel encontrado para processar.")
    else:
        print(f"\nEncontrados {len(excel_files)} ficheiro(s) para processar.\n")
        
        output_row = 2
        for excel_file in sorted(excel_files):
            print(f"A extrair de {excel_file.name}")
            process_excel_file(excel_file, output_ws, output_row)
            output_row += 1
    
    # Save output file
    output_wb.save(OUTPUT_FILE)
    print(f"\n{'=' * 60}")
    print(f"Dados extraídos e guardados em {OUTPUT_FILE}")
    print(f"{'=' * 60}")
    
    input("\nEnter para sair...")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nOperação cancelada pelo utilizador.")
    except Exception as e:
        print(f"\n\nErro fatal: {e}")
        input("\nEnter para sair...")
